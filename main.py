# -*- coding:utf-8 -*- 
import os
import sys
import datetime
from datetime import date, timedelta
import math
from openpyxl import Workbook # 엑셀
from openpyxl.styles import numbers
from pytz import timezone
# import urlparse
import telegram
import requests
import time
import ssl
import json
import re
import pymysql
import pymysql.cursors
from typing import List
from bs4 import BeautifulSoup
#from urllib.parse import urlparse
import urllib.parse as urlparse
import urllib.request

from requests import get  # to make GET request

# 텔레그램 봇
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Updater, CallbackQueryHandler, CommandHandler, MessageHandler , Filters

# 타임존 설정
# 한국(서울) 기준
datetime.datetime.now(timezone('Asia/Seoul'))
# UTC
datetime.datetime.now(timezone('UTC'))

# 참고 문서
# https://heodolf.tistory.com/75
# https://heodolf.tistory.com/76
# https://heodolf.tistory.com/77
# https://minmong.tistory.com/312

# 로직 설명
# 1. Main()-> 각 회사별 함수를 통해 반복 (추후 함수명 일괄 변경 예정)
#   - checkNewArticle -> parse -> downloadFile -> Send 
# 2. 연속키의 경우 현재 .key로 저장
#   - 추후 heroku db로 처리 예정(MySQL)
#   - DB연결이 안되는 경우, Key로 처리할수 있도록 예외처리 반영
# 3. 최초 조회되는 게시판 혹은 Key값이 없는 경우 메세지를 발송하지 않음.
# 4. 테스트와 운영을 구분하여 텔레그램 발송 채널 ID 구분 로직 추가
#   - 어떻게 구분지을지 생각해봐야함
# 5. 메시지 발송 방법 변경 (봇 to 사용자 -> 채널에 발송)

############텔레그램 전역변수###########
MSG = ''
############공용 상수############
# 선택 아이템
SELECT_ITEM = (
    "스크리닝 직접 입력모드",               # 0
    "마법공식",             # 1
    "마법공식 엑셀",           # 2
    "하나금융투자",          # 3
    "한양증권",              # 4
    "삼성증권",              # 5
    "교보증권"              # 6
)
##### 시스템 상수 #####
dir_now = os.path.dirname(os.path.abspath(__file__))  # real path to dirname

##### 엑셀 상수 #####
# 엑셀파일 읽기
load_wb  = ''
# 엑셀파일 쓰기
write_wb = Workbook()

# 초기 시트 삭제
write_wb.remove(write_wb['Sheet'])
# 이름이 있는 시트를 생성
write_ws = write_wb.create_sheet('Sheet1')

# Sheet1에다 입력
write_ws = write_wb.active
# 엑셀출력 상수
EXCEL_TITLE = ( # 엑셀은 인덱스가 아님! (순번1부터)
    "구분",                       # 1 
    "섹터&업종",                  # 2
    "종목명",                     # 3
    "전일종가(원)",               # 4
    "PER",                        # 5
    "Fwd PER",                     # 6
    "PBR",                       # 7
    "ROE",                       # 8
    "DPS Yield(%)",             # 9
    "OPM Yield(%)",             # 10
    "매출액(억원)",             # 11
    "영업이익(억원)",           # 12
    "당기순이익(억원)",         # 13
    "시가총액(억원)",          # 14
    "자본총계(억원)",           # 15
    "거래소",                   # 16
    "네이버 금융",              # 17
    "fnguide",                  # 18
    "기업개요"                  # 19 
)

# secret key
TELEGRTargetBOT_TOKEN_MAGIC_FORMULA_SECRET = ''
# 메시지 발송 ID 
chat_id = ''
# 퀀트 URL 변수
TARGET_URL = '' 
# 퀀트 엑셀 변수
TARGET_FILE = ''

# pymysql 변수
conn    = ''
cursor  = ''

# 이모지
EMOJI_FIRE = u'\U0001F525'
EMOJI_PICK = u'\U0001F449'

# 엑셀 출력을 위한 열 인덱스
nColIdx = 0 

strFileName = ''

data_selected = 0

NAVER_URL= 'https://finance.naver.com/item/main.nhn?code='
# JSON API 타입
# http://wise.thewm.co.kr/ASP/Screener/Screener1.asp?ud=#tabPaging 
# 의 산출 정보를 이용하여 종목 스크리닝 상세 정보를 생성 
# args[0] = data_selected, 
# args[1] = Target, 
# args[2] = chat_id
def MagicFormula_crowling(*args):
    global strFileName
    global TARGET_URL
    global TARGET_FILE
    global write_wb
    global data_selected

    print("***************MagicFormula_crowling********************")

    data_selected = 2
    today = date.today()
    yesterday = date.today() - timedelta(1)    
    DEFAULT_URL = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT='+ yesterday.strftime('%Y%m%d') +'&termCount=3&currentPage=1&orderKey=P1&orderDirect=D&jsonParam=%5B%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%2231%22%2C%22MIN_VAL%22%3A%220.01%22%2C%22MAX_VAL%22%3A%2219%22%2C%22Ogb%22%3A%222%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%223.00%22%2C%22MAX_VAL%22%3A%2220.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22P%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%2210.00%22%2C%22MAX_VAL%22%3A%221388%22%2C%22Ogb%22%3A%222%22%7D%5D'

    try: # 사용자의 입력값 확인 (args[0])
        if args[0] == 0 or args[0] == 2:
            print(args[0])
            print("0번모드 사용자 입력모드")
            TARGET_URL = str(args[1]).strip()
        elif args[0] == 1:
            print("1번모드")
            TARGET_URL = DEFAULT_URL 
            return  sendText("1.마법공식 종목받기는 준비중입니다. ㄷㄷ")
    except IndexError: # 사용자가 입력한 입력값 없을때
        print("스크리닝 URL에러로 기본값으로")
        TARGET_URL = DEFAULT_URL
        # TARGET_URL = 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp?market=0&industry=G0&size=0&workDT=' +  yesterday.strftime('%Y%m%d') +'&termCount=3&currentPage=1&orderKey=V1&orderDirect=A&jsonParam=%5B%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%22-3.93%22%2C%22MAX_VAL%22%3A%2222.89%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22P%22%2C%22SEQ%22%3A%221%22%2C%22MIN_VAL%22%3A%225.00%22%2C%22MAX_VAL%22%3A%2240.00%22%2C%22Ogb%22%3A%221%22%7D%2C%7B%22Group%22%3A%22V%22%2C%22SEQ%22%3A%2232%22%2C%22MIN_VAL%22%3A%221.00%22%2C%22MAX_VAL%22%3A%228%22%2C%22Ogb%22%3A%222%22%7D%5D'

    strUserReqDate = ''
    try: # 사용자의 입력 형식 확인 (args[1])
        strUserReqDate = args[1]
    except:
        sendText("입력한 URL 혹은 파일 형식이 올바르지 않습니다.")
        return True

    if "http:" in strUserReqDate : strTargetType = "URL"
    else : strTargetType = "EXCEL"

    if strTargetType == 'URL':
        try:
            workDt = TARGET_URL.find("&workDT=")
        except IndexError:
            return  sendText("입력한 스크리닝URL이 올바르지 않습니다.")

        
        if workDt < 0 : return # 입력하신 URL이 올바르지 않습니다.
        else: 
            if args[0]:
                userWorkdt = '&workDT=' + TARGET_URL[workDt+8:workDt+16]

        print('###URL 확인###')
        print(TARGET_URL)
        request = urllib.request.Request(TARGET_URL)
        #검색 요청 및 처리
        response = urllib.request.urlopen(request)
        rescode = response.getcode()
        if rescode != 200 :return sendText("(http://wise.thewm.co.kr)사이트 접속이 원활하지 않습니다. 잠시후 다시 시도해주세요.")

        CMP_PAGE_CNT = 10
        jres = json.loads(response.read().decode('utf-8'))
        print(jres)
        try:
            TOTAL_CMP_CNT = jres['sAllCnt']
        except KeyError: # While문으로 처리하기 (임시조치 )
            userWorkdt = '&workDT=' + TARGET_URL[workDt+8:workDt+16]
            TARGET_URL = TARGET_URL.strip().replace(userWorkdt, '&workDT=' + yesterday.strftime('%Y%m%d') )
            request = urllib.request.Request(TARGET_URL)
            #검색 요청 및 처리
            response = urllib.request.urlopen(request)
            rescode = response.getcode()
            if rescode != 200 :return sendText("(http://wise.thewm.co.kr)사이트 접속이 원활하지 않습니다. 잠시후 다시 시도해주세요.")


            CMP_PAGE_CNT = 10
            jres = json.loads(response.read().decode('utf-8'))
            TOTAL_CMP_CNT = jres['sAllCnt']
    else:
        print('excel')
        return 
    TOTAL_PAGE_CNT = math.ceil(TOTAL_CMP_CNT / CMP_PAGE_CNT) # 페이지 수 이므로 정수가 아닌 경우 +1
    
    print('한 페이지에 회사 수는', CMP_PAGE_CNT , "건 입니다.")
    print('조건에 부합하는 회사 수는 ',TOTAL_CMP_CNT, "건 입니다.")
    print(TOTAL_PAGE_CNT)
    print("VAL 값은 우측 상단의 값임")
    print("반복코드는 나중에")
    sendMessageText = "입력 받은 조건으로 집계를 시작합니다. \n"+"스크리닝 종목수는 "+ str(TOTAL_CMP_CNT) + " 개 입니다. \n 전체 산출시간은 " + "약 " +  str(math.ceil( (TOTAL_CMP_CNT * 1.5) / 60 )) + "분으로 예상됩니다."
    sendText("입력 받은 조건으로 집계를 시작합니다. \n"+"스크리닝 종목수는 "+ str(TOTAL_CMP_CNT) + " 개 입니다. \n 전체 산출시간은 " + "약 " +  str(math.ceil( (TOTAL_CMP_CNT * 1.5) / 60 )) + "분으로 예상됩니다." )

    strFileName = str(today)+'.txt'
    if data_selected == 0 or data_selected == 1: strFileName = str(today)+'.txt'
    else: strFileName = str(today)+'.xlsx'
    
    file = open( strFileName, 'w')    # hello.txt 파일을 쓰기 모드(w)로 열기. 파일 객체 반환

    print(args[0] , '0이면 일반, 2면 엑셀임!')    
    try:
        jres = jres['resultList']
    except:
        print("스크리닝 리스트를 받아오지 못함 + 서버가 정상이라 가정하고 workdt 공휴일 보정처리")

    nRowIdx = 0
    CURRENT_CMP_CNT = 0
    b20pYN = 1 # 미발송 = 1 
    b40pYN = 1 # 미발송 = 1
    b60pYN = 1 # 미발송 = 1
    b80pYN = 1 # 미발송 = 1
    for idx in range(1, TOTAL_PAGE_CNT+1):
        paging = 'currentPage='
        paging += str(idx)
        
        # print(TARGET_URL)
        request = urllib.request.Request(TARGET_URL)
        #검색 요청 및 처리
        response = urllib.request.urlopen(request)
        rescode = response.getcode()
        if rescode != 200 :return print("네이버 뉴스 접속이 원활하지 않습니다 ")
        jres = json.loads(response.read().decode('utf-8'))
        jres = jres['resultList']
        
        if data_selected == 0 or data_selected == 1:
            for r in jres:
                
                write = ''
                write += NAVER_URL + r['CMP_CD'] + '\t' +'종목명:' + r['CMP_NM_KOR'] + '\n'
                write += fnguide_parse(r['CMP_CD']) + '\n'
                print(write)
                file.write(write)      # 파일에 문자열 저장
                nRowCnt=+1
        else: ## 2
            excel_write_title()
            for r in jres:
                nRowIdx= nRowIdx + 1
                CURRENT_CMP_CNT = nRowIdx
                excel_write_row(r['CMP_CD'], nRowIdx)
                
        TARGET_URL = TARGET_URL.replace( 'currentPage='+ str(idx), 'currentPage='+ str(idx+1) )
        CURRENT_PROGRESS_PERCENT = int(CURRENT_CMP_CNT / TOTAL_CMP_CNT * 100)
        print("CURRENT_PROGRESS_PERCENT", CURRENT_PROGRESS_PERCENT, "CURRENT_CMP_CNT", CURRENT_CMP_CNT, "TOTAL_CMP_CNT", TOTAL_CMP_CNT)
        if CURRENT_PROGRESS_PERCENT in range(0, 19): pass
        elif CURRENT_PROGRESS_PERCENT in range(20, 39) and b20pYN: sendEditText(sendMessageText + "\n*20% 진행 되었습니다.*"); b20pYN = 0
        elif CURRENT_PROGRESS_PERCENT in range(40, 59) and b40pYN: sendEditText(sendMessageText + "\n*40% 진행 되었습니다.*"); b40pYN = 0
        elif CURRENT_PROGRESS_PERCENT in range(60, 79) and b60pYN: sendEditText(sendMessageText + "\n*60% 진행 되었습니다.*"); b60pYN = 0
        elif CURRENT_PROGRESS_PERCENT in range(80, 99) and b80pYN: sendEditText(sendMessageText + "\n*80% 진행 되었습니다.*"); b80pYN = 0
        

    if data_selected == 0 or data_selected == 1:
        file.close() # 파일 객체 닫기
    elif data_selected == 2:
        write_wb.save(strFileName)

    sendEditText(sendMessageText + "\n*완료 되었습니다!*")    
    sendDocument() # txt, excel 발송 통합
    sendText('/start 를 눌러 시작해보세요.')

    return True

def sendText(sendMessageText): # 가공없이 텍스트를 발송합니다.
    global chat_id
    global MSG

    print('sendText()')
    bot = telegram.Bot(token = TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET)

    # print(chat_id)
    # bot.sendDocument(chat_id = chat_id, text = sendMessageText)
    MSG = bot.sendMessage(chat_id = chat_id, text = sendMessageText, disable_web_page_preview = True, parse_mode = "Markdown")
    
    time.sleep(2) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)
def sendEditText(sendMessageText): # 가공없이 텍스트를 발송합니다.
    global chat_id
    global MSG

    print('sendEditText()')
    bot = telegram.Bot(token = TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET)

    # print(MSG)
    # bot.sendDocument(chat_id = chat_id, text = sendMessageText)
    MSG = bot.editMessageText(chat_id = chat_id, message_id = MSG.message_id , text = sendMessageText, disable_web_page_preview = True, parse_mode = "Markdown")
    
    time.sleep(2) # 모바일 알림을 받기 위해 8초 텀을 둠(loop 호출시)
def sendDocument(): # 가공없이 첨부파일을 발송합니다.
    global chat_id

    print('sendDocument()')
    bot = telegram.Bot(token = TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET)

    # print(chat_id)
    bot.sendDocument(chat_id = chat_id, document =  open( strFileName, 'rb'))

    # bot.sendMessage(chat_id = chat_id, text = sendMessageText, disable_web_page_preview = True, parse_mode = "Markdown")
def fnguide_parse(*args):

    pattern = ''
    CODE = ''
    for pattern in args:
        if len(pattern) > 0 :
            CODE =  pattern

    TARGET_URL = 'http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?MenuYn=Y&gicode=A'
    TARGET_URL += CODE
    # 005930
    webpage = requests.get(TARGET_URL, verify=False)

    # HTML parse
    soup = BeautifulSoup(webpage.content, "html.parser")
    data_cmp_nm = soup.select_one('#giName').text
    data_cmp_code = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > h2').text
    data_stxt1 = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > p > span.stxt.stxt1').text
    data_stxt2 = soup.select_one('#strMarketTxt').text
    data_세부업종 = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > p > span.stxt.stxt2').text.strip()
    data_영업이익률 = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(15) > td:nth-child(4)').text.strip()
    data_Per = soup.select_one('#corp_group2 > dl:nth-child(1) > dd').text
    data_Pbr = soup.select_one('#corp_group2 > dl:nth-child(4) > dd').text.strip()
    data_Roe = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(18) > td:nth-child(4)').text.strip()
    data_fwdPer = soup.select_one('#corp_group2 > dl:nth-child(2) > dd').text
    data_dividendYield = soup.select_one('#corp_group2 > dl:nth-child(5) > dd').text
    data_cmp_info = soup.select_one('#bizSummaryContent').text
    #data_ROE = soup.select_one('#svdMainGrid10D > table > tbody > tr:nth-child(7) > td:nth-child(2)')#.text

    
    r = ''
    r += TARGET_URL + '\n'
    r += '==============================================================' + '\n'
    r += '종목명: ' + data_cmp_nm                                + '\n'
    r += '종목코드: ' + data_cmp_code                            + '\n'
    r += '업종: ' + data_stxt1                                      + '\t' + data_stxt2 +" / "+ data_세부업종 +'\n'
    r += 'PER(FY0): ' + data_Per                                      + '\n'
    r += '12m FWD PER: ' + data_fwdPer                           + '\n'
    r += 'PBR(FY0): ' + data_Pbr                                      + '\n'
    r += 'ROE(FY0): ' + data_Roe                                      + '\n'
    r += 'OPM(%): ' + data_영업이익률                                      + '\n'
    r += '시가배당 수익률 '  + data_dividendYield                + '\n'
    r += '기업개요:' + data_cmp_info                 + '\n'
    r += '==============================================================' + '\n'
    #print('ROE', data_ROE)
    
    
    return r
def excel_write_title(*args):
    
    # 타이틀
    for idx in range(0, len(EXCEL_TITLE)):
        write_ws.cell(1, idx+1, EXCEL_TITLE[idx])

    write_ws.auto_filter.ref = "A1:S1"
    write_ws.freeze_panes = 'A2' # 첫번째 Row 틀고정(타이틀)
def excel_write_row(*args):
    nColIdx = 0    # 항목별 출력을 위한 출력 열 인덱스
    strIsuNo = str(args[0])
    nRowIdx  = int(args[1]) + 1 # 첫번째 레코드는 헤더를 쓰기 때문

    TARGET_URL = 'http://comp.fnguide.com/SVO2/ASP/SVD_Main.asp?MenuYn=Y&gicode='+ 'A'+ strIsuNo
    NAVER_URL= 'https://finance.naver.com/item/main.nhn?code=' + strIsuNo
    FNGUIDE_URL = TARGET_URL

    webpage = requests.get(TARGET_URL, verify=False)

    # HTML parse
    soup = BeautifulSoup(webpage.content, "html.parser")
    data_cmp_nm     = soup.select_one('#giName').text.strip()
    data_cmp_code   = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > h2').text.strip()
    data_price      =  soup.select_one('#svdMainGrid1 > table > tbody > tr.rwf > td:nth-child(2)').text.split("/")[0].strip()
    data_stxt1      = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > p > span.stxt.stxt1').text.strip()
    data_업종       = data_stxt1.strip()
    data_세부업종   = soup.select_one('#compBody > div.section.ul_corpinfo > div.corp_group1 > p > span.stxt.stxt2').text.strip()
    data_stxt2      = soup.select_one('#strMarketTxt').text.strip()
    data_영업이익률 = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(15) > td:nth-child(4)').text.strip()
    data_Per        = soup.select_one('#corp_group2 > dl:nth-child(1) > dd').text.strip()
    data_Pbr        = soup.select_one('#corp_group2 > dl:nth-child(4) > dd').text.strip()
    data_Roe        = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(18) > td:nth-child(4)').text.strip()
    data_Roa        = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(17) > td:nth-child(4)').text.strip()
    
    data_Dpsyield   = soup.select_one('#corp_group2 > dl:nth-child(5) > dd').text.strip().replace('%','')
    data_시가총액   = soup.select_one('#svdMainGrid1 > table > tbody > tr:nth-child(5) > td:nth-child(2)').text.strip()
    data_매출액     = soup.select_one('#highlight_B_A > table > tbody > tr:nth-child(1) > td:nth-child(4)').text.strip()
    data_영업이익   = soup.select_one('#highlight_B_A > table > tbody > tr:nth-child(2) > td:nth-child(4)').text.strip()
    data_당기순이익 = soup.select_one('#highlight_B_A > table > tbody > tr:nth-child(4) > td:nth-child(4)').text.strip()
    data_자본총계   = soup.select_one('#highlight_D_A > table > tbody > tr:nth-child(7) > td:nth-child(4)').text.strip()
    data_fwdPer     = soup.select_one('#corp_group2 > dl:nth-child(2) > dd').text.strip()
    data_dividendYield = soup.select_one('#corp_group2 > dl:nth-child(5) > dd').text.strip()
    data_cmp_info = soup.select_one('#bizSummaryContent').text.strip()
    data_거래소 = data_stxt1.strip()#data_stxt1.strip().split(" ")[0].split("\xa0\xa0")[1]
    
    # print(nRowIdx, data_cmp_nm)

    # 첫번째 열은 사용하지 않음
    data_업종 = str(data_업종).replace('KSE','').replace('KOSDAQ', '').replace('코스피', '').replace('코스닥','').strip()
    if len(data_업종) == 0 : data_업종 = str(data_세부업종).replace("FICS",'').strip()

    SetColIdx(2) # 열 인덱스 출력 시작점

    write_ws.cell(nRowIdx, GetColIdx(1), data_업종)
    write_ws.cell(nRowIdx, GetColIdx(1), data_cmp_nm)

    if data_price not in ('', '-'): data_price = float(data_price.replace(',',''))
    else: data_price = ''
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_price)
    cell.number_format = '#,##0'

    if data_Per not in ('', '-', 'N/A'): data_Per = float(data_Per.replace(',',''))
    else: data_Per = '-'
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_Per)
    cell.number_format = '#,##0.00'

    if data_fwdPer not in ('', '-', 'N/A'): data_fwdPer = float(data_fwdPer.replace(',',''))
    else: data_fwdPer = '-'
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_fwdPer)
    cell.number_format = '#,##0.00'

    if data_Pbr not in ('', '-'): data_Pbr = float(data_Pbr.replace(',',''))
    else: data_Pbr = ''
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_Pbr)
    cell.number_format = '#,##0.00'

    try:
        data_Roe = float(data_Roe.replace(',',''))
    except:
        data_Roe = ''
    #if data_Roe not in ('', '-') : data_Roe = float(data_Roe.replace(',',''))
    #else: data_Roe = ''
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_Roe)
    cell.number_format = '#,##0.00'

    if data_Dpsyield not in ('', '-'): data_Dpsyield = float(data_Dpsyield.replace(',',''))
    else: data_Dpsyield = ''
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_Dpsyield)
    cell.number_format = '#,##0.00'

    if data_영업이익률 not in ('', '-'): data_영업이익률 = float(data_영업이익률.replace(',',''))
    else: data_영업이익률 = ''
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_영업이익률)
    cell.number_format = '#,##0.00'

    if data_매출액 not in ('', '-'): data_매출액 = float(data_매출액.replace(',',''))
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_매출액)
    cell.number_format = '#,##0'

    if data_영업이익 not in ('', '-'): data_영업이익 = float(data_영업이익.replace(',',''))
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_영업이익)
    cell.number_format = '#,##0'

    if data_당기순이익 not in ('', '-'): data_당기순이익 = float(data_당기순이익.replace(',',''))
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_당기순이익)
    cell.number_format = '#,##0'

    if data_시가총액 not in ('', '-'): data_시가총액 = float(data_시가총액.replace(',',''))
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_시가총액)
    cell.number_format = '#,##0'

    if data_자본총계 not in ('', '-'): data_자본총계 = float(data_자본총계.replace(',',''))
    cell = write_ws.cell(nRowIdx, GetColIdx(1), data_자본총계)
    cell.number_format = '#,##0'
    

    if 'KOSDAQ' in data_거래소 :  data_거래소 = 'KOSDAQ'
    else: data_거래소 = 'KOSPI'
    write_ws.cell(nRowIdx, GetColIdx(1), data_거래소)
    
    # 네이버 링크
    write_ws.cell(nRowIdx, GetColIdx(0)).hyperlink = NAVER_URL
    write_ws.cell(nRowIdx, GetColIdx(0)).value =  data_cmp_nm+ '('+ data_cmp_code +')'
    write_ws.cell(nRowIdx, GetColIdx(1)).style = "Hyperlink"

    # FNGUIDE 링크
    write_ws.cell(nRowIdx, GetColIdx(0)).hyperlink = FNGUIDE_URL
    write_ws.cell(nRowIdx, GetColIdx(0)).value =  data_cmp_nm+ '('+ data_cmp_code +')'
    write_ws.cell(nRowIdx, GetColIdx(1)).style = "Hyperlink"

    write_ws.cell(nRowIdx, GetColIdx(1), str(data_cmp_info).strip())
def SetColIdx(*args):
    global nColIdx

    try:
        nColIdx = args[0]
    except:
        nColIdx = 0 

    return nColIdx
# ColIdx 반환과 인덱스 계산 함수
# 인자값에 만큼 인덱스를 증가 시킴
# 인자가 없는 경우 0으로 간주
# return 인자는 nCurColIdx (현재 인덱스)
def GetColIdx(*args):
    global nColIdx

    nCurColIdx = nColIdx
    try:
        nIncrIdx = args[0]
    except:
        nIncrIdx = 0 

    nColIdx = nColIdx + nIncrIdx
    
    return nCurColIdx
# 시간 및 날짜는 모두 한국 시간 (timezone('Asia/Seoul')) 으로 합니다.
def GetCurrentDate(*args):
    pattern = ''
    for pattern in args:
        print('pattern 입력값',pattern)
    
    time_now = str(datetime.datetime.now(timezone('Asia/Seoul')))[:19] # 밀리세컨즈 제거

    DATE = time_now[:10].strip()
    DATE_SPLIT = DATE.split("-")

    if pattern == '':
        DATE = time_now[:10].strip()
    elif pattern == 'YY' or pattern == 'yy':
        DATE = DATE_SPLIT[0][2:]
    elif pattern == 'YYYY' or pattern == 'yyyy':
        DATE = DATE_SPLIT[0]
    elif pattern == 'MM' or pattern == 'mm':
        DATE = DATE_SPLIT[1]
    elif pattern == 'DD' or pattern == 'dd':
        DATE = DATE_SPLIT[2]
    elif pattern == 'YYYY/HH/MM' or pattern == 'yyyy/hh/mm':
        DATE = DATE_SPLIT[0] + "/" + DATE_SPLIT[1] + "/" + DATE_SPLIT[2]
    elif pattern == 'YYYY-HH-MM' or pattern == 'yyyy-hh-mm':
        DATE = time_now[:10].strip()
    elif pattern == 'YY-HH-MM' or pattern == 'yy-hh-mm':
        DATE = time_now[2:10].strip()
    elif pattern == 'YYYYMMDD' or pattern == 'yyyymmdd':
        DATE = DATE_SPLIT[0] + DATE_SPLIT[1] + DATE_SPLIT[2]
    else:
        DATE = time_now[:10].strip()

    return DATE
def start(update, context):
    global chat_id
    
    chat_id = update.message.chat_id
    task_buttons =  [
        [ InlineKeyboardButton( '0. 스크리닝 직접 입력모드', callback_data=0 ) ],
        [ InlineKeyboardButton( '1.마법공식 종목받기(TTM PER 20배 이내, 배당 지급이력, ROE 10%↑ (PER내림차순 정렬)', callback_data=1 ) ],
        [ InlineKeyboardButton( '2.마법공식 엑셀', callback_data=2 ) ] ,
        [ InlineKeyboardButton( '3.준비중', callback_data=3 ) ] 
    ]
    
    reply_markup = InlineKeyboardMarkup( task_buttons )
    
    context.bot.send_message(
        chat_id=update.message.chat_id
        , text='작업을 선택해주세요.'
        , reply_markup=reply_markup
    )
def callback_get(update, context):
    global data_selected
    global chat_id 

    chat_id = update.callback_query.message.chat_id
    print("callback")
    data_selected = int(update.callback_query.data)
    print(data_selected)

    if data_selected == 0 or data_selected == 2:
        # 스크리닝 URL 안내 발송
        context.bot.edit_message_text(text="{}이(가) 선택되었습니다".format(SELECT_ITEM[data_selected]),
                                    chat_id=update.callback_query.message.chat_id,
                                    message_id=update.callback_query.message.message_id)
        context.bot.send_message(chat_id=update.callback_query.message.chat_id, text="가이드 링크 : " + 'https://www.notion.so/shinseunghoon/URL-9b91ddd9b409479ca9a0276d0c5a69be' + '\n' + '\n'+ '스크리닝 링크 : '+ 'http://wise.thewm.co.kr/ASP/Screener/Screener1.asp?ud=#tabPaging')
        context.bot.send_message(chat_id=update.callback_query.message.chat_id, text="가이드를 참조하여 스크리닝 URL을 입력하세요.")


    elif data_selected == 1:
        context.bot.edit_message_text(text="{}이(가) 선택되었습니다".format(SELECT_ITEM[data_selected]),
                                    chat_id=update.callback_query.message.message_id,
                                    message_id=update.callback_query.message.message_id)
        MagicFormula_crowling(1)
def get_screening_url(update, context):
    global chat_id 

    chat_id = update.message.chat_id
    if data_selected == 0 or data_selected == 2:
        inputURL = update.message.text
        # URL 형태가 아닌 경우 다시 입력을 받을 수 있는지 여부 확인
        if 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp' not in inputURL:
            context.bot.send_message(chat_id=update.message.chat_id, text="가이드를 참조하여 스크리닝 URL을 입력하세요.")
            

        if inputURL.find("&workDT=") < 0 :
            context.bot.send_message(chat_id=update.message.chat_id, text="스크리닝 URL을 재생성 해주세요.")
        
        URL = update.message.text
        MagicFormula_crowling(data_selected, URL, chat_id)          
def get_screening_excel(update, context):
    global chat_id 

    chat_id = update.message.chat_id
    file_id_short = update.message.document.file_id
    file_url = os.path.join(dir_now, update.message.document.file_name)
    file_name = update.message.document.file_name
    file_extension = update.message.document.file_name
    
    try:
        file_extension = file_extension.split(".")[1]
        # print(file_extension)
    except:
        update.message.reply_text('파일 형식이 올바르지 않습니다. 올바른 파일을 전송해주세요.')

    # print(file_extension)
    if  "xl" not in file_extension  : update.message.reply_text('엑셀 형식이 아닙니다. 올바른 파일을 전송해주세요.')
    elif file_extension == "xls" : Convert_xlsx(file_name)

    # 사용자 요청 파일 저장(서버 다운로드)
    context.bot.getFile(file_id_short).download(file_url)
    update.message.reply_text('전송한 종목 리스트로 집계를 시작합니다.')
    f = file_name
    if excel_read_file(f):
        pass
    else:
        return False
    # print(file_id_short, file_url, file_name)
    return 
    if data_selected == 0 or data_selected == 2:
        inputURL = update.message.text
        # URL 형태가 아닌 경우 다시 입력을 받을 수 있는지 여부 확인
        if 'http://wise.thewm.co.kr/ASP/Screener/data/Screener_Termtabledata.asp' not in inputURL:
            context.bot.send_message(chat_id=update.message.chat_id, text="가이드를 참조하여 스크리닝 URL을 입력하세요.")
            

        if inputURL.find("&workDT=") < 0 :
            context.bot.send_message(chat_id=update.message.chat_id, text="스크리닝 URL을 재생성 해주세요.")
        
        URL = update.message.text
        MagicFormula_crowling(data_selected, URL, chat_id)          
def Convert_xlsx(xls_file_path):
    xlsBook = xlrd.open_workbook(xls_file_path)
    workbook = openpyxl.Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = workbook.active if i == 0 else workbook.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                colvalue = xlsSheet.cell_value(row, col)
                print(colvalue)
                if isinstance(colvalue, str):
                    colvalue = colvalue.replace('', ' ', 3)		
	
                sheet.cell(row=row + 1, column=col + 1).value = colvalue
    
    return workbook
def excel_read_file(*args):
    global load_wb
    
    strExcelFilePath = str(args[0])
    print('strExcelFilePath', strExcelFilePath)
    #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    load_wb = openpyxl.load_workbook("./"+strExcelFilePath) #, data_only=True)
    # 
    # 
    
    # openpyxl.utils.exceptions.InvalidFileException: openpyxl does not support the old .xls file format, please use xlrd to read this file, or convert it to the more recent .xlsx file format.
    # https://soulmatt.tistory.com/entry/%ED%8C%8C%EC%9D%B4%EC%8D%AC%EC%97%90%EC%84%9C-xlrd-%EB%9D%BC%EC%9D%B4%EB%B8%8C%EB%9F%AC%EB%A6%AC-%EC%9D%B4%EC%9A%A9%ED%95%B4%EC%84%9C-xls-xlsx-%EB%B3%80%ED%99%98%ED%95%98%EA%B8%B0-How-to-convert-xls-to-xlsx-in-python
    # try:
    #     strExcelFilePath = str(args[0])
    #     print('strExcelFilePath', strExcelFilePath)
    #     #data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
    #     load_wb = openpyxl.load_workbook("./"+strExcelFilePath) #, data_only=True)
    # except:
    #     if len(strExcelFilePath): print(strExcelFilePath)
    #     else: print('엑셀 파일 경로를 확인 해주세요.')

    ws = load_wb.active
    print('sheetsname:',ws.title)
    # print(load_wb.sheetnames)
    for sheet_nm in load_wb.sheetnames: 
        print('*' * 100) 
        print('시트명:', sheet_nm) 
        sheet = load_wb[sheet_nm] 
        for row_data in sheet.iter_rows(min_row=1): # min_row는 시작 행을 지정 
            for cell in row_data: 
                print('[', cell.value, ']') 
            print('=' * 100) 

    load_wb.close()



    # #시트 이름으로 불러오기
    # load_ws = load_wb['Sheet1']
    
    # #셀 주소로 값 출력
    # print(load_ws['A1'].value)
    
    # #셀 좌표로 값 출력
    # print(load_ws.cell(1,2).value)

    return True
def GetSecretKey(*args):
    global TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET

    SECRETS = ''
    # print(os.getcwd())
    if os.path.isfile(os.path.join(os.getcwd(), 'secrets.json')): # 로컬 개발 환경
        with open("secrets.json") as f:
            SECRETS = json.loads(f.read())
        TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET     =   SECRETS['TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET']
        
    else: # 서버 배포 환경(heroku)
        TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET     =   os.environ.get('TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET')

def main():
    global chat_id
    global data_selected
    
    print('########Program Start Run########')
    
    GetSecretKey()

    updater = Updater( token=TELEGRAM_BOT_TOKEN_MAGIC_FORMULA_SECRET, use_context=True )

    # 버튼 UI dispatcher
    dispatcher = updater.dispatcher

    dispatcher.add_handler(CommandHandler("start", start))
    dispatcher.add_handler(CallbackQueryHandler(callback_get))          # updater.dispatcher.add_handler(CallbackQueryHandler(callback_get))
    updater.dispatcher.add_handler(MessageHandler(Filters.text, get_screening_url))
    updater.dispatcher.add_handler(MessageHandler(Filters.document, get_screening_excel))

    updater.start_polling()
    updater.idle()

if __name__ == "__main__":
	main()